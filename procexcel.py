# This is a cod for load from EXCEL
import pandas as pd
import os
from openpyxl import load_workbook 
import math

# каталог хранения EXCRL файлов
os.chdir("d:\_project\excel-repo")
# загрузить файл с данными, который хранится в каталоге
wb = load_workbook(filename = os.listdir('.')[0])
ws = wb.active
# список листов в файле EXCEL/ предполагаем что листов 3
sheet1 = wb['Лист1']
sheet2 = wb['Лист2']
sheet3 = wb['Лист3']

for i in range(2,7):
    for j in range(2, 7):
        print(i, j, ' -> ', sheet1.cell(row=i, column=j).value)


